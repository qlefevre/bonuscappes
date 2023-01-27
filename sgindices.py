from openpyxl import load_workbook
from io import BytesIO
import urllib.request
import base64
from datetime import date
from tempfile import NamedTemporaryFile


def handle(event, context):
    # Fichiers Excel
    srcWb = load_workbook_from_url(
        'https://bourse.societegenerale.fr/EmcWebApi/api/ProductSearch/Export?PageNum=1&ProductClassificationId=19&AssetTypeId=2&AssetTypeMenuId=35')
    # srcWb = load_workbook(path.join(sys.path[0], 'export.xlsx'))
    modWb = load_workbook_from_url(
        'https://raw.githubusercontent.com/qlefevre/bonuscappes/main/xlsx/modele.xlsx')
    # modWb = load_workbook(path.join(sys.path[0], 'modele.xlsx'))
    # Onglets Export
    srcWs = srcWb['EXPORT']
    modWs = modWb['EXPORT']
    # Copie les données de l'onglet Export du fichier source vers le modèle
    for row in srcWs:
        for cell in row:
            modWs[cell.coordinate].value = cell.value

    output = save_virtual_workbook(modWb)
    with open("Bonus_Cappes_SG_Indices_"+date.today().strftime("%Y%m%d")+".xlsx", "wb") as binary_file:
        # Write bytes to file
        binary_file.write(output)

    # return {
    #     "body":  base64.b64encode(output).decode('UTF-8'),
    #     "statusCode": 200,
    #     "isBase64Encoded": True,
    #     "headers": {
    #         "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #         "Content-Disposition": "attachment;filename=Bonus_Cappes_SG_Indices_"+date.today().strftime("%Y%m%d")+".xlsx"
    #     }
    # }


def load_workbook_from_url(url):
    # Charge un fichier Excel depuis url
    file = urllib.request.urlopen(url).read()
    return load_workbook(filename=BytesIO(file))


def save_virtual_workbook(workbook):
    # Charge le dernier jeu de données et l'importe dans le modèle
    with NamedTemporaryFile(delete=False) as tf:
        workbook.save(tf.name)
        in_memory = BytesIO(tf.read())
        return in_memory.getvalue()


if __name__ == '__main__':
    handle(None, None)
