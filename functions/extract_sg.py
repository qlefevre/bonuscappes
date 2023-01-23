from openpyxl import load_workbook
from os import path
from io import BytesIO
from io import BytesIO
from tempfile import NamedTemporaryFile
import sys
import urllib.request
import base64

# Charge un fichier Excel depuis url


def load_workbook_from_url(url):
    file = urllib.request.urlopen(url).read()
    return load_workbook(filename=BytesIO(file))


def save_virtual_workbook(workbook):
    with NamedTemporaryFile(delete=False) as tf:
        workbook.save(tf.name)
        in_memory = BytesIO(tf.read())
        return in_memory.getvalue()

# Charge le dernier jeu de données et l'importe dans le modèle


def handle(event, context):

    # Fichiers Excel
    srcWb = load_workbook_from_url(
        'https://bourse.societegenerale.fr/EmcWebApi/api/ProductSearch/Export?PageNum=1&ProductClassificationId=19&AssetTypeId=2&AssetTypeMenuId=35')
    # srcWb = load_workbook(path.join(sys.path[0], 'export.xlsx'))
    modWb = load_workbook_from_url(
        'https://raw.githubusercontent.com/qlefevre/bonuscappes/main/functions/modele.xlsx')
    # modWb = load_workbook(path.join(sys.path[0], 'modele.xlsx'))
    # Onglets Export
    srcWs = srcWb['EXPORT']
    modWs = modWb['EXPORT']
    # Copie les données de l'onglet Export du fichier source vers le modèle
    for row in srcWs:
        for cell in row:
            modWs[cell.coordinate].value = cell.value

    # modWb.save(path.join(sys.path[0], 'resultat.xlsx'))

    output = save_virtual_workbook(modWb)

    return {
        "body":  base64.b64encode(output).decode('UTF-8'),
        "statusCode": 200,
        "isBase64Encoded": True,
        "headers": {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": "attachment;filename=test.xlsx"
        }
    }


# main
if __name__ == '__main__':
    handle(None, None)
